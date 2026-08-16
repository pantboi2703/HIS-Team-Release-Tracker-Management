"""Excel export.

The release team still receives an .xlsx, and reproducing their layout exactly
is what kills the "we'll just keep using Excel" objection. So the export is not
a generic dump: it rebuilds the original sheet — metadata rows above the header,
the same column order — plus a stats sheet carrying the two pivots.

The stats are written as live COUNTIFS formulas, not computed values, so the
recipient's file behaves like the original one they are used to. XLOOKUP, FILTER
and UNIQUE are deliberately avoided: they break in LibreOffice and older Excel,
and this file gets opened on whatever is installed on the ward PC.
"""

from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..models.enums import Status

# The source column order, reproduced exactly.
COLUMNS = [
    ("RM", "rm", 10),
    ("Link", "link", 10),
    ("Tracker", "tracker", 14),
    ("Description", "subject", 62),
    ("Module", "module", 14),
    ("Assignee", "assignee_name", 16),
    ("Testing Status", "status_label", 16),
    ("Showstopper?", "showstopper_label", 17),
    ("Testing Remark", "remark", 60),
    ("Business Impact ", "business_impact", 20),
]

STATUS_LABEL = {
    "NOT_STARTED": "Not Started",
    "WIP": "WIP",
    "PASS": "Pass",
    "FAIL": "Fail",
    "PARTIAL_PASS": "Partial Pass",
    "RETEST": "Retest",
    "UNABLE_TO_TEST": "Unable to Test",
    "NOT_REPRODUCIBLE": "Not Reproducible",
}

INK = "16324F"
HEADER_FILL = PatternFill("solid", fgColor="F6F6F3")
THIN = Side(style="thin", color="E0DFD8")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _showstopper_label(value: Any) -> str:
    if value is True:
        return "Show stopper"
    if value is False:
        return "Not a Showstopper"
    return ""


def build_workbook(
    cycle: dict[str, Any],
    runs: list[dict[str, Any]],
    testers: list[dict[str, Any]],
) -> Workbook:
    wb = Workbook()
    ws = wb.active
    phase = cycle["phase"]
    ws.title = f"{phase} {cycle['build']} Testing"[:31]

    # --- metadata rows above the header, as in the source file ---
    ws.cell(1, 1, f"{phase} Product Version: {cycle['release']} Build: {cycle['build']}").font = Font(bold=True)
    ws.cell(2, 1, f"{phase} Testing Start")
    ws.cell(2, 4, cycle.get("start_date") or "")
    ws.cell(3, 1, f"{phase} Test End")
    ws.cell(3, 4, cycle.get("end_date") or "")

    header_row = 5
    for idx, (label, _key, width) in enumerate(COLUMNS, start=1):
        c = ws.cell(header_row, idx, label)
        c.font = Font(bold=True, color=INK)
        c.fill = HEADER_FILL
        c.border = BORDER
        ws.column_dimensions[get_column_letter(idx)].width = width

    for offset, run in enumerate(runs):
        row = header_row + 1 + offset
        values = {
            "rm": int(run["rm"]) if str(run["rm"]).isdigit() else run["rm"],
            "link": int(run["rm"]) if str(run["rm"]).isdigit() else run["rm"],
            "tracker": run.get("tracker", ""),
            "subject": run.get("subject", ""),
            "module": run.get("module", ""),
            "assignee_name": run.get("assignee_name") or "",
            "status_label": STATUS_LABEL.get(run["status"], run["status"]),
            "showstopper_label": _showstopper_label(run.get("showstopper")),
            "remark": run.get("remark") or "",
            "business_impact": run.get("business_impact") or "",
        }
        for idx, (_label, key, _w) in enumerate(COLUMNS, start=1):
            c = ws.cell(row, idx, values[key])
            c.border = BORDER
            c.alignment = Alignment(vertical="top", wrap_text=key in ("subject", "remark"))
        if run.get("redmine_url"):
            link_cell = ws.cell(row, 2)
            link_cell.hyperlink = run["redmine_url"]
            link_cell.font = Font(color="0563C1", underline="single")

    ws.freeze_panes = ws.cell(header_row + 1, 1)
    last_row = header_row + len(runs)

    _stats_sheet(wb, ws.title, header_row, last_row, testers, cycle)
    return wb


def _stats_sheet(
    wb: Workbook,
    data_sheet: str,
    header_row: int,
    last_row: int,
    testers: list[dict[str, Any]],
    cycle: dict[str, Any],
) -> None:
    """Status-wise and person-wise pivots, as live formulas.

    Quoting the sheet name matters: 'L2-A B123 Testing' contains a hyphen and a
    space, so an unquoted reference produces #REF! — which is exactly the state
    the current stats sheet in the real file is in.
    """
    ws = wb.create_sheet("Stats")
    sheet = f"'{data_sheet}'"
    status_col = get_column_letter(7)   # Testing Status
    person_col = get_column_letter(6)   # Assignee
    stopper_col = get_column_letter(8)  # Showstopper?

    status_range = f"{sheet}!${status_col}${header_row + 1}:${status_col}${max(last_row, header_row + 1)}"
    person_range = f"{sheet}!${person_col}${header_row + 1}:${person_col}${max(last_row, header_row + 1)}"
    stopper_range = f"{sheet}!${stopper_col}${header_row + 1}:${stopper_col}${max(last_row, header_row + 1)}"

    ws.cell(1, 1, f"Status wise summary — {cycle['name']}").font = Font(bold=True, color=INK)
    ws.cell(2, 1, "Status").font = Font(bold=True)
    ws.cell(2, 2, "#s").font = Font(bold=True)

    labels = [STATUS_LABEL[s.value] for s in Status]
    for i, label in enumerate(labels):
        r = 3 + i
        ws.cell(r, 1, label)
        ws.cell(r, 2, f'=COUNTIF({status_range},A{r})')
    total_row = 3 + len(labels)
    ws.cell(total_row, 1, "Total").font = Font(bold=True)
    ws.cell(total_row, 2, f"=SUM(B3:B{total_row - 1})").font = Font(bold=True)

    ws.cell(1, 5, f"User wise — {cycle['name']}").font = Font(bold=True, color=INK)
    ws.cell(2, 5, "Assigned").font = Font(bold=True)
    ws.cell(2, 6, "Assigned #s").font = Font(bold=True)
    for i, label in enumerate(labels):
        ws.cell(2, 7 + i, label).font = Font(bold=True)
    ws.cell(2, 7 + len(labels), "Showstoppers").font = Font(bold=True)

    for j, tester in enumerate(testers):
        r = 3 + j
        ws.cell(r, 5, tester["full_name"])
        ws.cell(r, 6, f"=COUNTIF({person_range},$E{r})")
        for i in range(len(labels)):
            col = 7 + i
            letter = get_column_letter(col)
            ws.cell(r, col, f"=COUNTIFS({person_range},$E{r},{status_range},{letter}$2)")
        ws.cell(
            r,
            7 + len(labels),
            f'=COUNTIFS({person_range},$E{r},{stopper_range},"Show stopper")',
        )

    tail = 3 + len(testers)
    ws.cell(tail, 5, "Total").font = Font(bold=True)
    for col in range(6, 8 + len(labels)):
        letter = get_column_letter(col)
        ws.cell(tail, col, f"=SUM({letter}3:{letter}{tail - 1})").font = Font(bold=True)

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["E"].width = 22
    for i in range(len(labels) + 1):
        ws.column_dimensions[get_column_letter(7 + i)].width = 15

    ws.cell(tail + 2, 1, "Counts are live COUNTIF formulas over the testing sheet, so editing a")
    ws.cell(tail + 3, 1, "status there updates this sheet, exactly like the original workbook.")
