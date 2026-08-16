"""The five-stage import pipeline.

    1 PARSE    read workbook, auto-detect header row, skip non-testing sheets
    2 MAP      headers -> canonical fields; unknown columns -> extra{}
    3 RESOLVE  normalise statuses; resolve assignee names via aliases;
               detect duplicate RM numbers within the sheet
    4 PREVIEW  STOP. Return a diff. Nothing written yet.
               >>> ADMIN CONFIRMS, MAPS NAMES, RESOLVES DUPLICATES <<<
    5 COMMIT   bulk upsert + run_history entries + save raw file to disk

Stage 4 is non-negotiable. A silent import that guesses wrong is worse than no
import at all.

The quirks handled here are all real, taken from the R18 workbook:
  * the header sits on row 5 under "L1 Product Version: 6.3.R18 Build: 111"
  * a second column repeats the RM number as a hyperlink and has no header
  * "Bugz", "Adt", "Partially Pass", "not reproduceable", trailing spaces
  * "Kamal " (trailing space) and "divitya" (lowercase) must resolve to people
  * RM 199385 appears twice on one sheet with two different descriptions
"""

import re
from datetime import date, datetime
from typing import Any

from openpyxl import load_workbook

from ..models.enums import (
    MODULE_MAP,
    SHOWSTOPPER_MAP,
    STATUS_MAP,
    TRACKER_MAP,
)

# Canonical field <- any of these header spellings (lowercased, stripped).
HEADER_ALIASES: dict[str, str] = {
    "rm": "rm",
    "#": "rm",
    "rm no": "rm",
    "rm number": "rm",
    "link": "link",
    "tracker": "tracker",
    "description": "subject",
    "subject": "subject",
    "module": "module",
    "assignee": "assignee",
    "assigned to": "assignee",
    "testing status": "status",
    "status": "status",
    "showstopper?": "showstopper",
    "showstopper": "showstopper",
    "show stopper": "showstopper",
    "testing remark": "remark",
    "remark": "remark",
    "remarks": "remark",
    "business impact": "business_impact",
}

MAX_HEADER_SCAN = 15

# A sheet with RM + Assignee is not necessarily a testing sheet: the raw Redmine
# export ("issues (45)") has both, but its Assignee column holds developers and
# it has no verdict on it. A testing sheet always carries at least one of these.
TESTING_MARKERS = {"testing status", "testing remark", "showstopper?", "showstopper", "show stopper"}


def _clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()[:10]
    return re.sub(r"\s+", " ", str(value)).strip()


def _strip_tags(text: str) -> str:
    """Real subjects are prefixed with routing tags: '<hpl><Sooraj> Please ...'."""
    return re.sub(r"\s+", " ", re.sub(r"<[^>]{0,40}>", "", text)).strip()


def find_header_row(ws) -> tuple[int | None, str | None]:
    """Never hardcode the header row number. Scan the first rows for one that has
    an RM (or #) column, an Assignee column, and at least one testing-specific
    column. Real workbooks contain 'Read me', pivot sheets and a raw Redmine
    export, and all of those must be skipped rather than half-imported.

    Returns (header_row, skip_reason)."""
    limit = min(ws.max_row, MAX_HEADER_SCAN)
    best_reason = "no RM column and no Assignee column"
    for row in range(1, limit + 1):
        labels = {
            _clean(ws.cell(row, col).value).lower()
            for col in range(1, min(ws.max_column, 30) + 1)
        }
        canon = {HEADER_ALIASES.get(lbl) for lbl in labels if lbl}
        if "rm" in canon and "assignee" in canon:
            if labels & TESTING_MARKERS:
                return row, None
            best_reason = "has RM and Assignee but no testing status or remark column"
    return None, best_reason


def count_data_rows(ws, header_row: int) -> int:
    """ws.max_row is inflated by stray formatting — the R18 sheets report ~995
    rows for 56 real ones. Count rows that actually carry an RM number."""
    n = 0
    rm_col = None
    for col in range(1, min(ws.max_column, 30) + 1):
        if HEADER_ALIASES.get(_clean(ws.cell(header_row, col).value).lower()) == "rm":
            rm_col = col
            break
    if rm_col is None:
        return 0
    for row in range(header_row + 1, ws.max_row + 1):
        if _clean(ws.cell(row, rm_col).value).isdigit():
            n += 1
    return n


def describe_sheet(ws, name: str) -> dict[str, Any]:
    header, reason = find_header_row(ws)
    return {
        "name": name,
        "header_row": header,
        "rows": count_data_rows(ws, header) if header else 0,
        "skipped": reason,
    }


def map_headers(ws, header_row: int) -> tuple[dict[str, int], dict[str, int]]:
    """Returns (canonical field -> column index, unknown header -> column index)."""
    known: dict[str, int] = {}
    extra: dict[str, int] = {}
    for col in range(1, min(ws.max_column, 40) + 1):
        raw = _clean(ws.cell(header_row, col).value)
        if not raw:
            continue
        field = HEADER_ALIASES.get(raw.lower())
        if field and field not in known:
            known[field] = col
        elif not field:
            extra[raw] = col
    return known, extra


def read_metadata(ws, header_row: int) -> dict[str, str | None]:
    """The rows above the header carry the release, build and start date, e.g.
    'L1 Product Version: 6.3.R18 Build: 111' then 'L1Testing Start | 2026-07-17'."""
    release = phase = build = start = None
    for row in range(1, header_row):
        joined = " ".join(
            _clean(ws.cell(row, col).value) for col in range(1, min(ws.max_column, 10) + 1)
        ).strip()
        if not joined:
            continue
        if m := re.search(r"(\d+\.\d+\.R\d+)", joined, re.I):
            release = m.group(1)
        if m := re.search(r"Build[:\s]*0*(\d+)", joined, re.I):
            build = f"B{m.group(1)}"
        if m := re.search(r"\b(L2-A|L2|L1|UAT)\b", joined, re.I):
            phase = phase or m.group(1).upper().replace("L2-A", "L2-A")
        if re.search(r"testing\s*start", joined, re.I):
            if m := re.search(r"(\d{4}-\d{2}-\d{2})", joined):
                start = m.group(1)
    return {"release": release, "phase": phase, "build": build, "start_date": start}


def list_sheets(path: str) -> list[dict[str, Any]]:
    wb = load_workbook(path, data_only=True)
    out = [describe_sheet(wb[name], name) for name in wb.sheetnames]
    wb.close()
    return out


def parse_sheet(path: str, sheet: str | None = None) -> dict[str, Any]:
    """Stages 1-3. Returns rows plus everything the preview screen needs."""
    wb = load_workbook(path, data_only=True)
    sheets = [describe_sheet(wb[name], name) for name in wb.sheetnames]

    chosen = None
    if sheet:
        chosen = next((s for s in sheets if s["name"] == sheet and s["header_row"]), None)
    if chosen is None:
        # Prefer the biggest testing sheet, so a 4-row leftover tab does not win.
        candidates = [s for s in sheets if s["header_row"]]
        chosen = max(candidates, key=lambda s: s["rows"], default=None)
    if chosen is None:
        wb.close()
        raise ValueError(
            "No sheet in this workbook has both an RM column and an Assignee column."
        )

    ws = wb[chosen["name"]]
    header_row = chosen["header_row"]
    known, extra_cols = map_headers(ws, header_row)
    metadata = read_metadata(ws, header_row)

    # Not every testing sheet repeats the product version — the R18 "L2-A
    # Testing" tab just says "In L2-A". Borrow the release from a sibling sheet
    # so step 4 opens pre-filled instead of blank. The admin can still correct it.
    if not metadata.get("release"):
        for other in sheets:
            if other["name"] == chosen["name"] or not other["header_row"]:
                continue
            sibling = read_metadata(wb[other["name"]], other["header_row"])
            if sibling.get("release"):
                metadata["release"] = sibling["release"]
                break

    rows: list[dict[str, Any]] = []
    warnings: list[dict[str, Any]] = []

    def cell(row: int, field: str) -> Any:
        col = known.get(field)
        return ws.cell(row, col).value if col else None

    for row in range(header_row + 1, ws.max_row + 1):
        rm_raw = cell(row, "rm")
        rm = _clean(rm_raw)
        if not rm:
            # A row with content but no RM number is never silently dropped —
            # the Patient Portal sheet has several of these.
            if _clean(cell(row, "subject")) or _clean(cell(row, "assignee")):
                warnings.append(
                    {
                        "row": row,
                        "what": "row has a description but no RM number",
                        "action": "row skipped, add it by hand if it is real",
                    }
                )
            continue
        if not rm.isdigit():
            # "Feature" appears in the RM column of the Patient Portal sheet.
            warnings.append({"row": row, "what": f'RM column reads "{rm}"', "action": "row skipped, add it by hand if needed"})
            continue
        if isinstance(rm_raw, str):
            warnings.append({"row": row, "what": f'RM number stored as text "{rm_raw}"', "action": f"read as {rm}"})

        status_raw = _clean(cell(row, "status"))
        status = STATUS_MAP.get(status_raw.lower())
        if status is None:
            # Anything unmatched surfaces as a warning. Never silently dropped.
            warnings.append({"row": row, "what": f'unknown status "{status_raw}"', "action": "set to Not started"})
            status = "NOT_STARTED"
        elif status_raw == "":
            warnings.append({"row": row, "what": "empty status", "action": "set to Not started"})

        ss_raw = _clean(cell(row, "showstopper"))
        showstopper = SHOWSTOPPER_MAP.get(ss_raw.lower(), "UNKNOWN")
        if showstopper == "UNKNOWN":
            warnings.append({"row": row, "what": f'unknown showstopper "{ss_raw}"', "action": "left unset"})
            showstopper = None
        elif ss_raw.lower() in ("y", "n"):
            warnings.append({"row": row, "what": f'showstopper column reads "{ss_raw}"', "action": f"read as {'yes' if showstopper else 'no'}"})

        module_raw = _clean(cell(row, "module"))
        module = MODULE_MAP.get(module_raw.lower(), module_raw)
        if module_raw and module != module_raw:
            warnings.append({"row": row, "what": f'module "{module_raw}"', "action": f"matched to {module}"})

        tracker_raw = _clean(cell(row, "tracker"))
        tracker = TRACKER_MAP.get(tracker_raw.lower(), tracker_raw or "Bug")

        assignee_raw = _clean(cell(row, "assignee"))

        rows.append(
            {
                "row": row,
                "rm": rm,
                "tracker": tracker,
                "subject": _strip_tags(_clean(cell(row, "subject"))),
                "module": module,
                "assignee_raw": assignee_raw,
                "status": status,
                "showstopper": showstopper,
                "remark": _clean(cell(row, "remark")),
                "business_impact": _clean(cell(row, "business_impact")) or None,
                "extra": {name: _clean(ws.cell(row, col).value) for name, col in extra_cols.items()},
            }
        )

    wb.close()

    # Duplicate RM numbers inside one sheet. Real, not hypothetical: in the R18
    # file RM 199385 sits on two rows with two different descriptions, because
    # one ticket sometimes covers two distinct pieces of work.
    seen: dict[str, list[dict[str, Any]]] = {}
    for r in rows:
        seen.setdefault(r["rm"], []).append(r)
    duplicates = [
        {
            "rm": rm,
            "rows": [
                {"row": d["row"], "subject": d["subject"], "module": d["module"], "assignee_raw": d["assignee_raw"]}
                for d in dupes
            ],
        }
        for rm, dupes in seen.items()
        if len(dupes) > 1
    ]

    return {
        "sheet": chosen["name"],
        "header_row": header_row,
        "sheets": sheets,
        "metadata": metadata,
        "rows": rows,
        "warnings": warnings,
        "duplicates": duplicates,
        "unknown_columns": sorted(extra_cols),
    }


async def resolve_assignees(db, rows: list[dict[str, Any]]) -> dict[str, Any]:
    """Alias matching, case-insensitive and whitespace-trimmed.

    Without an alias table you hand-fix names on every import forever, because
    the sheet writes first names only.
    """
    users = await db.users.find({}).to_list(None)
    lookup: dict[str, dict[str, Any]] = {}
    for u in users:
        for alias in [u["username"], u["full_name"], *(u.get("aliases") or [])]:
            key = str(alias).strip().lower()
            if key:
                lookup.setdefault(key, u)

    counts: dict[str, int] = {}
    for r in rows:
        counts[r["assignee_raw"]] = counts.get(r["assignee_raw"], 0) + 1

    resolved, unknown = [], []
    for raw, n in sorted(counts.items(), key=lambda kv: -kv[1]):
        if not raw:
            continue  # blank assignee is allowed and lands in the Unassigned bucket
        user = lookup.get(raw.strip().lower())
        if user:
            how = "exact alias" if raw in (user.get("aliases") or []) else "alias, normalised"
            resolved.append(
                {"raw": raw, "user_id": str(user["_id"]), "person": user["full_name"], "rows": n, "how": how}
            )
        else:
            unknown.append({"raw": raw, "rows": n})
    return {"resolved": resolved, "unknown": unknown}


def build_preview(parsed: dict[str, Any], names: dict[str, Any], existing_rms: set[str]) -> dict[str, Any]:
    """Stage 4 payload. Warnings inform; unresolved names and duplicate RM
    numbers block the import."""
    rows = parsed["rows"]
    unknown_raw = {u["raw"] for u in names["unknown"]}
    new = sum(1 for r in rows if r["rm"] not in existing_rms)

    return {
        "sheet": parsed["sheet"],
        "header_row": parsed["header_row"],
        "sheets": parsed["sheets"],
        "metadata": parsed["metadata"],
        "row_count": len(rows),
        "counts": {
            "new": new,
            "existing": len(rows) - new,
            "warnings": len(parsed["warnings"]),
            "unknown_names": len(names["unknown"]),
            "duplicate_rm": len(parsed["duplicates"]),
        },
        "rows": [
            {
                "row": r["row"],
                "rm": r["rm"],
                "subject": r["subject"],
                "module": r["module"],
                "assignee_raw": r["assignee_raw"],
                "status": r["status"],
                "assignee_unknown": r["assignee_raw"] in unknown_raw,
            }
            for r in rows[:200]
        ],
        "warnings": parsed["warnings"][:200],
        "duplicates": parsed["duplicates"],
        "unknown_assignees": names["unknown"],
        "resolved_assignees": names["resolved"],
        "unknown_columns": parsed["unknown_columns"],
        "blocked": bool(names["unknown"] or parsed["duplicates"]),
    }
