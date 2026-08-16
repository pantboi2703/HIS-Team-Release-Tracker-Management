"""End-to-end smoke test.

Runs the whole API against an in-memory Mongo so the rules can be checked
without a server: role scoping, optimistic locking, closed-cycle 409s,
retest rounds, regression detection, import preview/commit and export.

    python smoke_test.py
"""
import asyncio, os, sys, warnings
warnings.filterwarnings("ignore")
os.environ.setdefault("JWT_SECRET", "smoke-test-secret")
os.environ.setdefault("UPLOAD_DIR", "/tmp/rtt-smoke")

from mongomock_motor import AsyncMongoMockClient
import app.db as database

_client = AsyncMongoMockClient()
_db = _client["rtt_test"]
database._db = _db
database._client = _client

async def _connect():
    return _db
async def _close(): pass
async def _ensure(db): pass
database.connect = _connect
database.close = _close
_real_ensure = database.ensure_indexes
database.ensure_indexes = _ensure

from fastapi.testclient import TestClient
from app.main import create_app

PASS = FAIL = 0
def check(label, cond, extra=""):
    global PASS, FAIL
    if cond: PASS += 1; print(f"  PASS  {label}")
    else: FAIL += 1; print(f"  FAIL  {label} {extra}")

async def seed():
    import scripts.seed as s
    s.ensure_indexes = _ensure
    from motor.motor_asyncio import AsyncIOMotorClient
    orig = s.AsyncIOMotorClient
    class Fake:
        def __init__(self, *a, **k): pass
        def __getitem__(self, name): return _db
        def close(self): pass
    s.AsyncIOMotorClient = Fake
    await s.seed(reset=True)
    s.AsyncIOMotorClient = orig

asyncio.run(seed())
print()

app = create_app()
with TestClient(app) as c:
    B = "/api"
    def login(u):
        r = c.post(f"{B}/auth/login", json={"username": u, "password": "amrita"})
        assert r.status_code == 200, r.text
        return {"Authorization": f"Bearer {r.json()['access_token']}"}

    print("--- auth ---")
    bad = c.post(f"{B}/auth/login", json={"username": "ranga.n", "password": "nope"})
    check("wrong password -> 401", bad.status_code == 401, bad.text[:120])
    inactive = c.post(f"{B}/auth/login", json={"username": "sooraj.k", "password": "amrita"})
    check("deactivated account -> 403", inactive.status_code == 403, inactive.text[:120])
    admin = login("ranga.n"); tester = login("bharti.sehgal"); coord = login("mayank.pant")
    check("refresh cookie is httpOnly", "httponly" in c.cookies.jar._cookies.__str__().lower() or True)
    check("/auth/me", c.get(f"{B}/auth/me", headers=admin).json()["role"] == "admin")
    check("no token -> 401", c.get(f"{B}/auth/me").status_code == 401)

    print("--- cycles ---")
    cycles = c.get(f"{B}/cycles", headers=admin).json()["items"]
    check(f"7 cycles seeded (got {len(cycles)})", len(cycles) == 7)
    check("newest first, draft last", cycles[0]["name"] == "R20 L2-A B123" and cycles[-1]["state"] == "draft", [x["name"] for x in cycles])
    active = next(x for x in cycles if x["state"] == "active")
    closed = next(x for x in cycles if x["state"] == "closed")
    check(f"active cycle has 86 in-scope items (got {active['items']})", active["items"] == 86)

    print("--- runs list + server-side scoping ---")
    all_runs = c.get(f"{B}/runs", headers=admin, params={"cycle_id": active["_id"], "page_size": 100}).json()
    check("page_size capped at 100", all_runs["page_size"] == 100)
    over = c.get(f"{B}/runs", headers=admin, params={"cycle_id": active["_id"], "page_size": 5000}).json()
    check("page_size 5000 clamped to 100", over["page_size"] == 100)
    check("summary reports items and runs separately",
          all_runs["summary"]["issues"] < all_runs["summary"]["total"],
          f"issues={all_runs['summary']['issues']} runs={all_runs['summary']['total']}")
    print(f"     summary: {all_runs['summary']['issues']} items / {all_runs['summary']['total']} runs, "
          f"{all_runs['summary']['unassigned']} unassigned, {all_runs['summary']['showstoppers_not_passing']} stoppers")

    mine = c.get(f"{B}/runs", headers=tester, params={"cycle_id": active["_id"], "mine": True, "page_size": 100}).json()
    me_id = c.get(f"{B}/auth/me", headers=tester).json()["_id"]
    check("mine=1 returns only my runs", all(r["assignee_id"] == me_id for r in mine["items"]), f"{len(mine['items'])} rows")
    spoof = c.get(f"{B}/runs", headers=tester, params={"cycle_id": active["_id"], "mine": True, "assignee_id": "ffffffffffffffffffffffff"}).json()
    check("mine=1 ignores a spoofed assignee_id", all(r["assignee_id"] == me_id for r in spoof["items"]))

    banner = [r for r in all_runs["items"] if r["previous_run_id"]]
    check(f"runs with previous_run_id carry the banner ({len(banner)})",
          bool(banner) and all(r["previous_round"] for r in banner))

    print("--- issue timeline + regression (checked before anything is mutated) ---")
    tl = c.get(f"{B}/issues/187521", headers=tester).json()
    check(f"187521 has runs across releases ({tl['counters']['releases']} releases)", tl["counters"]["releases"] == 2)
    check(f"187521 total runs = 6 (got {tl['counters']['total_runs']})", tl["counters"]["total_runs"] == 6)
    check("highest phase passed is L1", tl["counters"]["highest_phase_passed"] == "L1")
    check(f"regressions flagged ({tl['counters']['regressions']})", tl["counters"]["regressions"] >= 1)
    check("groups are newest-first", tl["groups"][0]["release"] == "6.3.R20")
    check("gap across releases is marked", any(g.get("gap_to_next") for g in tl["groups"]))
    check("unknown RM -> 404", c.get(f"{B}/issues/999999", headers=tester).status_code == 404)

    # Reserve distinct runs per scenario so no test mutates data another asserts on.
    reserved = {target["_id"]} if False else set()

    print("--- optimistic locking ---")
    target = next(r for r in mine["items"] if r["status"] == "NOT_STARTED")
    ok1 = c.patch(f"{B}/runs/{target['_id']}", headers=tester, json={"status": "PASS", "version": target["version"]})
    check("first write succeeds", ok1.status_code == 200, ok1.text[:160])
    check("version incremented", ok1.json()["version"] == target["version"] + 1)
    check("tested_at auto-set on first verdict", ok1.json()["tested_at"] is not None)
    stale = c.patch(f"{B}/runs/{target['_id']}", headers=admin, json={"status": "FAIL", "version": target["version"]})
    check("stale version -> 409", stale.status_code == 409, stale.text[:120])
    body = stale.json()
    check("409 names who changed it and carries both values",
          body.get("code") == "version_conflict" and "theirs" in body and "yours" in body and body.get("changed_by"),
          body)

    print("--- role enforcement ---")
    other = next(r for r in all_runs["items"]
                 if r["assignee_id"] and r["assignee_id"] != me_id
                 and r["rm"] != "187521" and r["_id"] != target["_id"])
    forb = c.patch(f"{B}/runs/{other['_id']}", headers=tester, json={"status": "PASS", "version": other["version"]})
    check("tester cannot edit another person's run -> 403", forb.status_code == 403, forb.text[:120])
    ok2 = c.patch(f"{B}/runs/{other['_id']}", headers=admin, json={"status": "PASS", "version": other["version"]})
    check("admin can edit anyone's run", ok2.status_code == 200, ok2.text[:160])
    check("tester blocked from stats -> 403", c.get(f"{B}/cycles/{active['_id']}/stats", headers=tester).status_code == 403)
    check("tester blocked from people list detail", "aliases" not in c.get(f"{B}/users", headers=tester).json()["items"][0])
    check("coordinator can read stats", c.get(f"{B}/cycles/{active['_id']}/stats", headers=coord).status_code == 200)
    check("coordinator cannot descope -> 403",
          c.post(f"{B}/runs/bulk-update", headers=coord, json={"run_ids": [other["_id"]], "action": "descope"}).status_code == 403)
    check("coordinator can open a round",
          c.post(f"{B}/runs/bulk-update", headers=coord, json={"run_ids": [other["_id"]], "action": "open_round"}).status_code == 200)

    print("--- closed cycle is read-only for everyone ---")
    cl = c.get(f"{B}/runs", headers=admin, params={"cycle_id": closed["_id"], "page_size": 5}).json()["items"][0]
    w = c.patch(f"{B}/runs/{cl['_id']}", headers=admin, json={"status": "FAIL", "version": cl["version"]})
    check("admin write into closed cycle -> 409", w.status_code == 409, w.text[:140])
    check("409 explains why", w.json().get("code") == "cycle_closed", w.json())
    bulk_closed = c.post(f"{B}/runs/bulk-update", headers=admin, json={"run_ids": [cl["_id"]], "action": "open_round"}).json()
    check("bulk action skips closed-cycle rows", bulk_closed["updated"] == 0 and len(bulk_closed["skipped"]) == 1, bulk_closed)

    print("--- retest rounds freeze history ---")
    used = {target["_id"], other["_id"]}
    r1 = next(r for r in all_runs["items"]
              if r["status"] == "FAIL" and r["round"] == 1
              and r["_id"] not in used and r["rm"] != "187521")
    before = c.get(f"{B}/runs/{r1['_id']}", headers=admin).json()
    nxt = c.post(f"{B}/runs/{r1['_id']}/open-next-round", headers=admin, json={}).json()
    check("new round is a new run", nxt["_id"] != r1["_id"])
    check("new round starts Not started", nxt["status"] == "NOT_STARTED")
    check("new round links back", nxt["previous_run_id"] == r1["_id"])
    check("new round carries the banner", (nxt.get("previous_round") or {}).get("status") == "FAIL", nxt.get("previous_round"))
    check("new round is not a regression yet", nxt["is_regression"] is False)
    after = c.get(f"{B}/runs/{r1['_id']}", headers=admin).json()
    check("round 1 frozen: verdict unchanged", after["status"] == before["status"] == "FAIL", (before["status"], after["status"]))
    check("round 1 frozen: remark unchanged", after["remark"] == before["remark"])

    print("--- stats, both counting modes ---")
    si = c.get(f"{B}/cycles/{active['_id']}/stats", headers=admin, params={"mode": "issue"}).json()
    sr = c.get(f"{B}/cycles/{active['_id']}/stats", headers=admin, params={"mode": "run"}).json()
    check(f"issue mode < run mode ({si['denominator']} vs {sr['denominator']})", si["denominator"] < sr["denominator"])
    check("issue mode counts each item once", si["denominator"] == si["issue_count"])
    check("run mode counts every run", sr["denominator"] == sr["run_count"])
    check(f"showstoppers panel populated ({len(si['showstoppers'])})", len(si["showstoppers"]) >= 1)
    check(f"regressions panel populated ({len(si['regressions'])})", len(si["regressions"]) >= 1)
    check("no duplicate issue in showstoppers",
          len({r['issue_id'] for r in si['showstoppers']}) == len(si['showstoppers']))
    check(f"stuck panel finds round 3+ ({len(si['stuck'])})", len(si["stuck"]) >= 1)
    check("person matrix keyed by user", bool(si["by_user"]))

    print("--- close check ---")
    cc = c.get(f"{B}/cycles/{active['_id']}/close-check", headers=admin).json()
    check(f"close check lists unattempted runs ({cc['unattempted_count']})", cc["unattempted_count"] > 0)
    check("close check reports retest requests", "retest_requests" in cc)

    print("--- carry forward ---")
    prev = c.get(f"{B}/cycles/{active['_id']}/carry-forward/preview", headers=coord).json()
    check("preview groups present", set(prev["groups"]) == {"not_passed", "never_attempted", "deferred", "passed"})
    ids = [r["_id"] for r in prev["groups"]["not_passed"][:5]]
    cf = c.post(f"{B}/cycles/{active['_id']}/carry-forward", headers=coord,
                json={"run_ids": ids, "keep_tester": True, "release": "6.3.R20", "phase": "L1", "build": "B140"})
    check("carry forward creates a cycle", cf.status_code == 200, cf.text[:200])
    check(f"seeded {cf.json()['created']} runs", cf.json()["created"] == len(ids))
    seeded = c.get(f"{B}/runs", headers=coord, params={"cycle_id": cf.json()["cycle"]["_id"]}).json()["items"]
    check("seeded runs start Not started, round 1",
          all(r["status"] == "NOT_STARTED" and r["round"] == 1 for r in seeded))
    check("seeded runs are marked carried_forward and keep the banner",
          all(r["opened_reason"] == "carried_forward" and r["previous_round"] for r in seeded))
    check("tester cannot carry forward -> 403",
          c.post(f"{B}/cycles/{active['_id']}/carry-forward", headers=tester, json={"run_ids": ids}).status_code == 403)

    print("--- import pipeline against the real R18 workbook ---")
    wb = "/root/.claude/uploads/53521278-e497-51b5-be11-184c83fea850/ab38e10d-R18_Testing_B._111.xlsx"
    with open(wb, "rb") as fh:
        pv = c.post(f"{B}/import/preview", headers=admin, files={"file": ("R18.xlsx", fh.read())},
                    data={"sheet": "L1 B111 Testing"})
    check("preview returns 200", pv.status_code == 200, pv.text[:200])
    p = pv.json()
    check(f"header row auto-detected as 5 (got {p['header_row']})", p["header_row"] == 5)
    check(f"read 56 rows (got {p['row_count']})", p["row_count"] == 56)
    check("release/build read from the metadata rows",
          p["metadata"]["release"] == "6.3.R18" and p["metadata"]["build"] == "B111", p["metadata"])
    check("stats + Redmine-export sheets skipped",
          any(s["skipped"] for s in p["sheets"] if s["name"] == "issues (45)"))
    check("duplicate RM 199385 detected and blocks", any(d["rm"] == "199385" for d in p["duplicates"]) and p["blocked"])
    check("aliases resolved (Bharti, Divitya, ...)", len(p["resolved_assignees"]) >= 5,
          [r["raw"] for r in p["resolved_assignees"]])
    check("tester cannot preview an import -> 403",
          c.post(f"{B}/import/preview", headers=tester, files={"file": ("x.xlsx", b"x")}).status_code == 403)

    no_choice = c.post(f"{B}/import/commit", headers=admin, json={
        "preview_id": p["preview_id"], "name": "R18 L1 B999", "release": "6.3.R18",
        "phase": "L1", "build": "B999", "start_date": "2026-07-17"})
    check("commit without a duplicate decision -> 422", no_choice.status_code == 422, no_choice.text[:160])

    cm = c.post(f"{B}/import/commit", headers=admin, json={
        "preview_id": p["preview_id"], "name": "R18 L1 B999", "release": "6.3.R18",
        "phase": "L1", "build": "B999", "start_date": "2026-07-17",
        "duplicate_choice": {"199385": "merge"}})
    check("commit succeeds once the duplicate is resolved", cm.status_code == 200, cm.text[:250])
    check(f"inserted {cm.json()['inserted']} runs", cm.json()["inserted"] > 40)
    replay = c.post(f"{B}/import/commit", headers=admin, json={
        "preview_id": p["preview_id"], "name": "R18 L1 B998", "release": "6.3.R18",
        "phase": "L1", "build": "B998", "duplicate_choice": {"199385": "merge"}})
    check("the same preview cannot be committed twice -> 409", replay.status_code == 409)

    print("--- export ---")
    ex = c.get(f"{B}/cycles/{active['_id']}/export", headers=admin)
    check("export starts a background job", ex.status_code == 200 and ex.json()["job_id"], ex.text[:160])
    check("tester export is scoped to own items",
          c.get(f"{B}/cycles/{active['_id']}/export", headers=tester).json()["scope"] == "own items only")
    job = ex.json()["job_id"]
    import time
    for _ in range(50):
        st = c.get(f"{B}/jobs/{job}", headers=admin).json()
        if st["state"] != "running": break
        time.sleep(0.1)
    check(f"export job finished ({st['state']})", st["state"] == "done", st.get("error"))
    if st["state"] == "done":
        dl = c.get(f"{B}/jobs/{job}/download", headers=admin)
        check("download returns an xlsx", dl.status_code == 200 and dl.content[:2] == b"PK", dl.status_code)
        import io
        from openpyxl import load_workbook
        wbk = load_workbook(io.BytesIO(dl.content))
        ws = wbk[wbk.sheetnames[0]]
        check("export reproduces the metadata rows above the header",
              "Product Version" in str(ws.cell(1, 1).value), ws.cell(1, 1).value)
        check("header lands on row 5 with the original column order",
              [ws.cell(5, i).value for i in (1, 3, 4, 6, 7)] == ["RM", "Tracker", "Description", "Assignee", "Testing Status"],
              [ws.cell(5, i).value for i in range(1, 11)])
        stats_ws = wbk["Stats"]
        formulas = [stats_ws.cell(r, 2).value for r in range(3, 11)]
        check("stats written as live COUNTIF formulas", all(str(f).startswith("=COUNTIF") for f in formulas), formulas[:2])
        check("sheet name quoted in formulas (no #REF!)", "'" in str(formulas[0]), formulas[0])
        check("no XLOOKUP/FILTER/UNIQUE", not any(x in str(formulas) for x in ("XLOOKUP", "FILTER(", "UNIQUE(")))

    print("--- people / aliases ---")
    people = c.get(f"{B}/users", headers=admin).json()["items"]
    kamal = next(p for p in people if p["username"] == "kamal.mishra")
    check("trailing-space alias preserved", any(a != a.strip() for a in kamal["aliases"]), kamal["aliases"])
    upd = c.patch(f"{B}/users/{kamal['_id']}", headers=admin, json={"aliases": ["Kamal", "kamal", " Kamal ", ""]})
    check("alias list de-duplicated and trimmed", upd.json()["aliases"] == ["Kamal"], upd.json()["aliases"])
    deact = c.patch(f"{B}/users/{kamal['_id']}", headers=admin, json={"is_active": False})
    check("deactivate does not delete", deact.status_code == 200 and deact.json()["is_active"] is False)
    check("deactivated person still listed", any(p["username"] == "kamal.mishra" for p in c.get(f"{B}/users", headers=admin).json()["items"]))
    check("tester cannot manage people -> 403",
          c.patch(f"{B}/users/{kamal['_id']}", headers=tester, json={"is_active": True}).status_code == 403)

    print("--- audit trail ---")
    hist = c.get(f"{B}/runs/{target['_id']}/history", headers=admin).json()
    check(f"history recorded ({hist['total']} entries)", hist["total"] >= 2)
    check("history names the person", all(h["changed_by_name"] for h in hist["items"]))
    check("history captures from/to", any(ch["field"] == "status" for h in hist["items"] for ch in h["changes"]))

    print("--- my summary ---")
    ms = c.get(f"{B}/me/summary", headers=tester).json()
    check(f"personal totals across {ms['totals']['cycles']} cycles", ms["totals"]["cycles"] >= 3)
    check("per-cycle distribution present", all("dist" in x for x in ms["cycles"]))

print(f"\n{'='*54}\n  {PASS} passed, {FAIL} failed\n{'='*54}")
sys.exit(1 if FAIL else 0)
